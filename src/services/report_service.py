"""Excel report generation for experiments.

Builds an .xlsx workbook containing:
  1. Experiment / film / equipment metadata block.
  2. A per-image table of scratch indices.
  3. A bar chart (histogram) of the scratch indices below the table.

The report is built purely from an ``ExperimentRead`` schema (film, config and
scratch_results are already loaded), so no extra DB access is needed.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from ..schemas.experiment import ExperimentRead

_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="4F6BED")
_LABEL_FONT = Font(bold=True)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2E7D32")


def _section_title(ws: Worksheet, row: int, text: str) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    return row + 1


def _kv(ws: Worksheet, row: int, label: str, value: Any) -> int:
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = _LABEL_FONT
    ws.cell(row=row, column=2, value="—" if value is None else value)
    return row + 1


def build_experiment_report_xlsx(experiment: ExperimentRead) -> bytes:
    """Build an Excel report for a single experiment and return its bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18

    title = ws.cell(
        row=1,
        column=1,
        value=f"Отчёт по эксперименту: {experiment.name or '—'}",
    )
    title.font = _TITLE_FONT

    row = 3
    row = _section_title(ws, row, "Эксперимент")
    row = _kv(ws, row, "Название", experiment.name)
    date_value = (
        experiment.date.replace(tzinfo=None) if experiment.date is not None else None
    )
    row = _kv(ws, row, "Дата", date_value)
    if date_value is not None:
        ws.cell(row=row - 1, column=2).number_format = "DD.MM.YYYY HH:MM"
    row = _kv(ws, row, "Вес образца, г", experiment.weight)  # noqa: RUF001
    row = _kv(ws, row, "Тканевая подложка", "Да" if experiment.has_fabric else "Нет")

    row += 1
    row = _section_title(ws, row, "Плёнка")
    film = experiment.film
    row = _kv(ws, row, "Название", film.name if film else None)
    row = _kv(ws, row, "Покрытие", film.coating_name if film else None)
    row = _kv(
        ws, row, "Толщина покрытия, мкм", film.coating_thickness if film else None
    )

    row += 1
    row = _section_title(ws, row, "Оборудование")
    config = experiment.config
    row = _kv(ws, row, "Конфигурация", config.name if config else None)
    row = _kv(ws, row, "Тип головки", config.head_type if config else None)
    row = _kv(ws, row, "Описание", config.description if config else None)

    # --- Indices table -------------------------------------------------
    row += 1
    table_header_row = row
    headers = ["Проходы", "Индекс царапины", "Кол-во пикселей"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    results = sorted(
        experiment.scratch_results or [],
        key=lambda r: (r.get("passes") if r.get("passes") is not None else 0),
    )
    data_start_row = row
    for result in results:
        ws.cell(row=row, column=1, value=result.get("passes") or 0)
        index_cell = ws.cell(row=row, column=2, value=result.get("scratch_index"))
        index_cell.number_format = "0.0000"
        ws.cell(row=row, column=3, value=result.get("total_pixels"))
        row += 1
    data_end_row = row - 1

    if results:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Гистограмма индексов царапины"
        chart.y_axis.title = "Индекс царапины"
        chart.x_axis.title = "Проходы"
        chart.legend = None

        data = Reference(
            ws,
            min_col=2,
            min_row=table_header_row,
            max_row=data_end_row,
        )
        cats = Reference(
            ws,
            min_col=1,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 18
        ws.add_chart(chart, f"A{data_end_row + 3}")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
